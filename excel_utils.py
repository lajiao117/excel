#!/usr/bin/env python
# encoding: utf-8
"""
@version: ??
@author: 
@software: PyCharm
@file: excel_utils.py
@time: 2018/8/1 15:29

"""
import os
import io
import time
import xlrd
import xlwt
import xlsxwriter
from urllib import parse
from flask import send_file, make_response
import mimetypes
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import time

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def export_template(filename, fields):
    # 导出模板xls文件
    book = xlwt.Workbook()
    sheet = book.add_sheet('Sheet1')  # 创建一个sheet
    for i in range(0, len(fields)):
        field = fields[i]
        sheet.write(0, i, field)


def export_data(filename, fields, data, names=None, sheet='Sheet1'):
    # fields 为list data为dict

    fp = io.BytesIO()
    book = xlsxwriter.Workbook(fp, {'in_memory': True})
    worksheet = book.add_worksheet(sheet)
    # 表头格式
    format1 = book.add_format(
        {'bold': True, 'font_color': 'black', 'font_size': 13, 'align': 'left', 'font_name': u'宋体'})
    # 表头外格式
    format2 = book.add_format({'font_color': 'black', 'font_size': 9, 'align': 'left', 'font_name': u'宋体'})
    # A列列宽设置能更好的显示
    worksheet.set_column("A:F", 20)
    # book = xlwt.Workbook()
    # worksheet = book.add_sheet('Sheet1')  # 创建一个sheet
    # 插入第一行表头标题
    if names:
        for i in range(0, len(names)):
            name = names[i]
            worksheet.write(0, i, name, format1)
    else:
        for i in range(0, len(fields)):
            field = fields[i]
            worksheet.write(0, i, field, format2)
    # 从第二行开始插入数据
    for i in range(len(data)):
        item = data[i]
        for j in range(len(fields)):
            field = fields[j]
            worksheet.write(i + 1, j, item[field])
    book.close()
    fp.seek(0)
    # print(filename,'////////////////////')
    name = parse.quote(filename)
    # print(fp.getvalue(),'-------------------')
    return send_file(fp, attachment_filename='%s.xlsx' % name, as_attachment=True)
    # mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def openxl_export_data(data, filename, sheet="sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for index, i in enumerate(data):
        for data_index, value in enumerate(i):
            ws.cell(row=index+1, column=data_index + 1, value=value)
    obj = save_virtual_workbook(wb)
    res = make_response(obj)
    # filename = '换行%s.xlsx' % int(time.time())
    mime_type = mimetypes.guess_type(filename)[0]
    res.headers['Content-type'] = mime_type
    res.headers['Content-Disposition'] = 'attachment; filename={}'.format(filename.encode().decode('latin-1'))
    # return send_file(obj, attachment_filename='%s.xlsx' % 'test', as_attachment=True)
    return res



